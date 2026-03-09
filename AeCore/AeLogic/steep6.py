import os
import shutil
from datetime import datetime
import pandas as pd
from aelogging import info, error
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def export_final_data_steep6(builder_instance, carpeta_salida, orden_trabajo):
    """
        Coge toda la información procesada y crea los excels finales,
        aplicando los estilos y copiando las fotos a sus carpetas.
    """
    info("Paso 6: Exportando archivos xls (Búsqueda Directa por ID Interno)")
    try:
        df = builder_instance.product.data_dt.copy()

        if df.empty or not carpeta_salida:
            error("Faltan datos o carpeta de salida.")
            return

        fecha_actual = datetime.now().strftime("%d/%m/%Y")
        prefijo_circuito = next(
            (os.path.basename(ruta).split('_')[0] for ruta in builder_instance.product.source_paths.values() if
             os.path.basename(ruta).split('_')[0].isdigit()), "")
        col_circ = next((c for c in df.columns if 'circuito' in str(c).lower()), None)
        col_vano = next((c for c in df.columns if 'vano' in str(c).lower() and 'long' not in str(c).lower()), None)

        if not prefijo_circuito:
            prefijo_circuito = ''.join(filter(str.isdigit, str(df[col_circ].iloc[0])))[:4] if col_circ else "0000"

        # generamos el el id unico
        df = df.reset_index(drop=True)
        df['ID'] = df.apply(
            lambda fila: f"1{prefijo_circuito}{str(fila.name).zfill(10 - len(f'1{prefijo_circuito}'))}",
            axis=1
        )

        col_ut = 'Ubicación Técnica'
        df[col_ut] = df.apply(
            lambda fila: str(fila.get(col_vano, '')).strip() if not str(fila.get(col_ut, '')).strip() or str(
                fila.get(col_ut, '')).strip().lower() in ['nan', 'none'] else str(fila.get(col_ut, '')).strip(), axis=1)

        # redondeamos a 2 decimales
        cols_a_redondear = ['x', 'y', 'z', 'LONGITUD_VANO', 'DIST_REGLA', 'Distancia_Final', 'CUMPLE']
        for col_red in cols_a_redondear:
            if col_red in df.columns:
                df[col_red] = pd.to_numeric(df[col_red], errors='coerce').round(2)

        # ordenamos de menor a mayor las anomalias por su ID
        df = df.sort_values(by='ID', ascending=True)

        nombre_circuito = str(df[col_circ].iloc[0]).strip() if col_circ else "CIRCUITO"

        #separacion de tablas
        es_vegetacion_mask = df['Tipo_Anomalia'].astype(str).str.lower().str.contains('vegetaci')
        df_infra = df[~es_vegetacion_mask].copy()
        df_vege = df[es_vegetacion_mask].copy()

        tipos_data = {
            'Infraestructura': (df_infra, [
                f"{nombre_circuito}_AnomaliasInfraestructuras",
                f"{nombre_circuito}_AnomaliasInfraestructuras_ANEXOX_LITE",
                f"{nombre_circuito}_AnomaliasInfraestructuras2026"
            ]),
            'Vegetación': (df_vege, [
                f"{nombre_circuito}_AnomaliasVegetacion",
                f"{nombre_circuito}_AnomaliasVegetacion_ANEXOX_LITE"
            ])
        }

        for tipo, (df_tipo, nombres_archivos) in tipos_data.items():
            if df_tipo.empty: continue

            ruta_fotos_tipo = os.path.join(carpeta_salida, tipo)
            os.makedirs(ruta_fotos_tipo, exist_ok=True)
            df_tipo['Evidencia_Grafica_Ruta'] = ""

            # buscamos foto por su id interno
            for index, fila in df_tipo.iterrows():
                # cogemos el ID original del informe
                id_interno = str(fila.get('ID_Original_Informe', '')).split('.')[0].strip()

                # cogemos la hipotesis de donde venía
                hipotesis_origen = str(fila.get('Hipotesis_Seleccionada', fila.get('hipotesis_origen', ''))).strip()

                if not id_interno or id_interno == 'nan' or id_interno == 'None':
                    continue

                archivo_real = None
                ruta_pictures_encontrada = None

                # 3. Buscamos directamente en las carpetas de origen
                for nombre_hoja, ruta_xls in builder_instance.product.source_paths.items():
                    # Buscamos en la carpeta que coincida con su hipótesis
                    if hipotesis_origen.upper() in nombre_hoja.upper() or nombre_hoja.upper() in hipotesis_origen.upper():
                        carpeta_origen = os.path.dirname(ruta_xls)
                        ruta_pictures = os.path.join(carpeta_origen, "pictures")

                        if os.path.exists(ruta_pictures):
                            # si la carpeta pictures existe, buscamos la imagen
                            for f in os.listdir(ruta_pictures):
                                if f.lower().endswith(('.jpg', '.jpeg', '.png')):
                                    nombre_sin_ext = os.path.splitext(f)[0]
                                    if nombre_sin_ext == id_interno or f.startswith(f"{id_interno}-") or f.startswith(
                                            f"{id_interno}_"):
                                        archivo_real = f
                                        ruta_pictures_encontrada = ruta_pictures
                                        break
                        if archivo_real:
                            break

                # si encontramos la foto, la copiamos y le ponemos el Super Nombre
                if archivo_real and ruta_pictures_encontrada:
                    id_nuevo_global = str(fila.get('ID', id_interno))
                    siglas_tipo = "VT" if tipo == 'Vegetación' else "IN"
                    hip_limpia = hipotesis_origen.replace(' ', '').replace('-', '').upper()
                    lin_limpia = nombre_circuito.replace('-', '').upper()
                    vano_limpio = str(fila.get(col_vano, 'VANO')).strip().replace('-', '_')

                    #
                    nuevo_nombre = f"{id_nuevo_global}_{siglas_tipo}_{hip_limpia}_{lin_limpia}_{vano_limpio}_{id_interno}{os.path.splitext(archivo_real)[1]}"

                    ruta_destino = os.path.join(ruta_fotos_tipo, nuevo_nombre)
                    shutil.copy2(os.path.join(ruta_pictures_encontrada, archivo_real), ruta_destino)
                    df_tipo.at[index, 'Evidencia_Grafica_Ruta'] = f"{tipo}/{nuevo_nombre}"

            #openpyxl
            for nombre_base in nombres_archivos:
                ruta_temp_xlsx = os.path.join(carpeta_salida,
                                              f"{orden_trabajo}_{nombre_base}.xlsx" if orden_trabajo else f"{nombre_base}.xlsx")
                ruta_final_xls = os.path.join(carpeta_salida,
                                              f"{orden_trabajo}_{nombre_base}.xls" if orden_trabajo else f"{nombre_base}.xls")

                if "LITE" in nombre_base:
                    cols_lite = ['DMR', 'huso', 'x', 'y', 'z', 'ID', 'Grado_Final', 'CIRCUITO', 'VANO', 'APOYO_INI',
                                 'APOYO_FIN', 'TENSION', 'LONGITUD_VANO', 'TIPO_ANO', 'REGLAMENTO', 'DIST_REGLA',
                                 'Distancia_Final', 'CUMPLE', 'NOTAS']
                    df_lite = df_tipo.reindex(columns=cols_lite)
                    df_lite.to_excel(ruta_temp_xlsx, index=False, startrow=3, header=False, engine='openpyxl')

                    wb = load_workbook(ruta_temp_xlsx)
                    ws = wb.active

                    ama = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                        bottom=Side(style='thin'))

                    alin_vertical = Alignment(horizontal='center', vertical='bottom', textRotation=90, wrapText=True)
                    alin_horizontal = Alignment(horizontal='center', vertical='center', wrapText=True)
                    alin_datos = Alignment(horizontal='center', vertical='center')

                    alin_coordenadas = Alignment(horizontal='center', vertical='center', textRotation=90, wrapText=True)

                    ws.merge_cells('C1:E1')
                    ws['C1'] = "UTM de las Anomalías"
                    ws['C1'].alignment = alin_horizontal
                    ws.merge_cells('C2:E2')
                    ws['C2'] = "Coordenadas"
                    ws['C2'].alignment = alin_coordenadas
                    ws['C3'] = "UTM (X)"
                    ws['C3'].alignment = alin_horizontal
                    ws['D3'] = "UTM (Y)"
                    ws['D3'].alignment = alin_horizontal
                    ws['E3'] = "UTM (Z)"
                    ws['E3'].alignment = alin_horizontal

                    titulos_verticales = {
                        1: "DMR", 2: "Huso", 6: "Nº de anomalía", 7: "Grado de anomalía",
                        8: "Nemónico circuito", 9: "Vano", 10: "Nº apoyo (a) inicial",
                        11: "Nº apoyo (a) final", 12: "Tensión (kV)", 13: "Longitud vano",
                        14: "Tipo de anomalía", 15: "Se aplica el reglamento", 16: "Distancia de reglamento",
                        17: "Mínima distancia absoluta de (h) a la trayectoria del cable",
                        18: "Cumple/incumple por +/- (m)", 19: "Notas"
                    }

                    for col_idx, texto in titulos_verticales.items():
                        letra = get_column_letter(col_idx)
                        ws.merge_cells(f'{letra}1:{letra}3')
                        ws[f'{letra}1'] = texto
                        ws[f'{letra}1'].alignment = alin_vertical

                    # Pintar cabecera
                    for fila in range(1, 4):
                        for col in range(1, 20):
                            celda = ws.cell(row=fila, column=col)
                            celda.fill = ama
                            celda.border = borde_fino

                    # Centrar los datos y poner cuadrícula a las filas de abajo
                    for r_idx in range(4, ws.max_row + 1):
                        for c_idx in range(1, 20):
                            celda = ws.cell(row=r_idx, column=c_idx)
                            celda.border = borde_fino
                            celda.alignment = alin_datos

                    ws.row_dimensions[1].height = 25
                    ws.row_dimensions[2].height = 140
                    ws.row_dimensions[3].height = 20

                    for i in range(1, 20):
                        if i in [3, 4, 5]:
                            ws.column_dimensions[get_column_letter(i)].width = 14
                        else:
                            ws.column_dimensions[get_column_letter(i)].width = 7

                    wb.save(ruta_temp_xlsx)
                    wb.close()

                elif "2026" in nombre_base:

                    df_ext = pd.DataFrame()

                    df_ext['Nº Anomalia SAP'] = df_tipo['ID']
                    df_ext['Ubicación técnica'] = df_tipo['Ubicación Técnica']
                    df_ext['Conjunto'] = df_tipo.get('Conjunto', "")
                    df_ext['Texto Ampliado'] = ""
                    df_ext['Repercusión'] = df_tipo.get('Grado_Final', 0)
                    df_ext['Reglamento de Aplicación'] = df_tipo.get('REGLAMENTO', "")
                    df_ext['Temperatura max de diseño de la linea'] = df_tipo.get('TEMP_MAX', "")
                    codigos = df_tipo.get('Codigo', "").fillna("")
                    descripciones = df_tipo.get('Descripción', "").fillna("")
                    df_ext['Código caracteristico - Descripción'] = codigos.astype(str) + " - " + descripciones.astype(
                        str)
                    df_ext['Distancia al primer apoyo del vano'] = df_tipo.get('DIST_APOYO', "")
                    df_ext['Fase/CT en la que se encuentra el incumplimiento'] = 3
                    df_ext['X'] = df_tipo.get('x', "")
                    df_ext['Y'] = df_tipo.get('y', "")
                    df_ext['HUSO'] = df_tipo.get('huso', "")
                    df_ext['Hipótesis climática de incumplimiento'] = df_tipo.get('Hipotesis_Seleccionada', "")
                    df_ext['Distancia reglamentaria considerada'] = df_tipo.get('DIST_REGLA', "")
                    df_ext['Distancia en el punto de incumplimiento'] = df_tipo.get('Distancia_Final', "")
                    df_ext.to_excel(ruta_temp_xlsx, index=False, engine='openpyxl')

                    wb = load_workbook(ruta_temp_xlsx)
                    ws = wb.active
                    ws.cell(row=1, column=17, value="Evidencia Gráfica")
                    link_font = Font(color="0000FF", underline="single")
                    for r_idx, evi in enumerate(df_tipo['Evidencia_Grafica_Ruta'], 2):

                        if evi:
                            c = ws.cell(row=r_idx, column=17, value="Imagen")
                            c.hyperlink = evi
                            c.font = link_font

                    for i in range(1, 18):
                        ws.column_dimensions[get_column_letter(i)].width = 22

                    wb.save(ruta_temp_xlsx)
                    wb.close()

                else:

                    df_n = pd.DataFrame()
                    df_n['ID'] = df_tipo['ID']
                    df_n['UT'] = df_tipo['Ubicación Técnica']
                    df_n['DESCRIPCIÓN'] = df_tipo.get('Descripción', "")
                    df_n['TEXTO AMPLIADO'] = ""
                    df_n['REPERCUSIÓN'] = df_tipo.get('Grado_Final', 0)
                    df_n['FECHA INICIO AVISO'] = fecha_actual
                    df_n['ORDEN ORIGEN'] = str(orden_trabajo)
                    df_n['FECHA INICIO AVERÍA'] = fecha_actual
                    df_n['OPERACIÓN'] = df_tipo.get('Operacion', "")

                    df_n.to_excel(ruta_temp_xlsx, index=False, engine='openpyxl')
                    wb = load_workbook(ruta_temp_xlsx)
                    ws = wb.active
                    ws.cell(row=1, column=10, value="EVIDENCIA GRÁFICA")
                    link_font = Font(color="0000FF", underline="single")
                    for r_idx, evi in enumerate(df_tipo['Evidencia_Grafica_Ruta'], 2):
                        if evi:
                            c = ws.cell(row=r_idx, column=10, value="Imagen")
                            c.hyperlink = evi
                            c.font = link_font
                    for i in range(1, 11): ws.column_dimensions[get_column_letter(i)].width = 22
                    wb.save(ruta_temp_xlsx)
                    wb.close()

                if os.path.exists(ruta_final_xls): os.remove(ruta_final_xls)
                os.rename(ruta_temp_xlsx, ruta_final_xls)
                info(f"Creado archivo: {nombre_base}")

    except Exception as e:
        error(f"Fallo al exportar: {e}")
        raise