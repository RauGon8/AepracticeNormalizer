import xlrd
import os
import shutil
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from aelogging import info, error
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


column_alias = {
    "ID_ANOMALIA": ["Nº de anomalía", "ID", "Anomalia_ID", "Número"],
    "GRADO": ["Grado", "Nivel", "Severidad", "Grado de anomalía"],
    "DISTANCIA": ["Distancia", "Metros", "Dist", "Mínima distancia absoluta de (h) a la trayectoria del cable"],
    "VANO": ["Vano", "Tramo", "Span"],
    "DENOMINACION_UT": ["Denominación UT", "Denominacion UT", "Denominacion", "Denominación"],
    "UBICACION_TECNICA": ["Ubicación Tecnica", "Ubicacion Tecnica", "Ubicación Técnica", "UT"],
    "x": ["X", "Coordenada X", "UTM (X)"],
    "y": ["Y", "Coordenada Y", "UTM (Y)"],
    "z": ["Z", "Coordenada Z", "UTM (Z)"],
    "huso": ["Huso", "Zona"],
    "CIRCUITO": ["Circuito", "Nombre de Circuito", "Nemónico circuito", "Línea"],
    "LONGITUD_VANO": ["Longitud vano", "Longitud", "Distancia Vano"],
    "DMR": ["DMR"],
    "APOYO_INI": ["Nº apoyo (a) inicial", "Apoyo Inicial"],
    "APOYO_FIN": ["Nº apoyo (b) final", "Nº apoyo (a) final", "Apoyo Final"],
    "TENSION": ["Tensión (kV)", "Tensión"],
    "TIPO_ANO": ["Tipo de anomalía"],
    "REGLAMENTO": ["Se aplica el reglamento", "Reglamento"],
    "DIST_REGLA": ["Distancia de reglamento"],
    "CUMPLE": [
        "Cumple/incumple por +/- (m)",
        "Cumple/incumple por +/-",
        "+/- distancia vertical (w) del cable a la cota de (h)",
        "+/- distancia horizontal (v) del punto (h) a la vertical del cable",
        "'+/- distancia vertical (w) del cable a la cota de (h)",
        "'+/- distancia horizontal (v) del punto (h) a la vertical del cable"
    ],
    "NOTAS": ["Notas y/o aclaraciones", "Notas", "Observaciones"],
    "TEMP_MAX": [
        "Temperatura máxima distinta",
        "Temperatura maxima distinta",
        "Temp ºC",
        "Temp. ºC",
        "Temperatura max de diseño de la linea",
        "Temperatura",
        "Temperatura máxima a aplicar"
    ],
    "FASE": ["Fase/CT", "Fase/CT en la que se encuentra el incumplimiento"],
    "DIST_APOYO": [
        "Distancia al apoyo más cercano",
        "Distancia al apoyo mas cercano",
        "Distancia al\napoyo  más\ncercano",
        "Distancia al apoyo  más cercano",
        "Distancia al apoyo",
        "Distancia al primer apoyo del vano"
    ]
}

def normalizar_texto(valor):
    """
        Limpia un texto quitándole espacios, tabulaciones y saltos de línea,
        y lo pasa a minúsculas para que sea fácil compararlo con otros textos.

        :param valor: El texto original que queremos limpiar.
        :return: El texto limpio en formato cadena.
        """
    if pd.isna(valor):
        return ""
    return str(valor).replace(" ","").replace("\t","").replace("\n","").strip().lower()

#funcion para comparar números con margen de error (Tolerancia)
def es_numero_parecido(val1, val2, tolerancia=0.5):
    def es_numero_parecido(val1, val2, tolerancia=0.5):
        """
        Compara dos números para ver si son casi iguales, dejando un pequeño
        margen de error (tolerancia) por si hay diferencias de decimales.

        :param val1: El primer número a comparar.
        :param val2: El segundo número a comparar.
        :param tolerancia: La diferencia máxima permitida entre los dos números (por defecto 0.5).
        :return: Verdadero (True) si son parecidos, Falso (False) si son muy distintos.
        """
    try:
        return abs(float(val1) - float(val2)) <= tolerancia
    except:
        return False

#reutilizo mi metodo para leer filas en los xls
def localizar_posiciones_cabeceras(hoja_lectura):
    """
    Escanea las primeras filas de un Excel como un radar para buscar
    los títulos de las columnas, usando nuestra lista de alias.

    :param hoja_lectura: La pestaña del archivo Excel que estamos escaneando.
    :return: Un diccionario con el nombre de la columna y sus coordenadas exactas (fila, columna).
    """
    posiciones={}
    #recorremos las primeras 20 filas

    for indice_fila in range (min(20,hoja_lectura.nrows)):
        #revisamos las columnas que tengan los datos de ncols
        for indice_columna in range (hoja_lectura.ncols):
            #cogemos el valor de cada celda, lo pasamos a str y luego le quitamos los espacios con el strip
            valor_celda=str(hoja_lectura.cell_value(indice_fila,indice_columna)).strip()

            #revisamos si el valor de la celda coincide con algun alias del diccionario
            for clave, lista_de_alias in column_alias.items():
                #comprobamos si encuentra una coincidencia y lo añade a posiciones, si encuentra x por ejemplo
                # no vuelve a buscarlo porque ya lo obtuvo (if clave not in posiciones)
                if clave not in posiciones:

                    #comparamos el valor con mi lista de column alias
                    if any(alias.lower()==valor_celda.lower()
                           for alias in lista_de_alias):
                        posiciones[clave]=(indice_fila,indice_columna)
    return posiciones

#lee informes y guarda datos normalizados
def extract_map_and_ids_steep1(builder_instance, rutas_hipotesis: list):
    """
    :param builder_instance:
    :param rutas_hipotesis:
    :return:
    """
    info("Procesando y normalizando identificadores de forma continua...")
    all_tables = []
    contador_global = 0

    for ruta_carpeta in rutas_hipotesis:
        nombre_hoja = os.path.basename(ruta_carpeta)
        archivos = [
            archivo for archivo in os.listdir(ruta_carpeta) if archivo.endswith('.xls')
        ]

        if not archivos: continue

        ruta_xls = os.path.join(ruta_carpeta, archivos[0])
        builder_instance.product.source_paths[nombre_hoja] = ruta_xls

        libro = xlrd.open_workbook(ruta_xls)
        hoja = libro.sheet_by_index(0)
        coords = localizar_posiciones_cabeceras(hoja)

        if "ID_ANOMALIA" in coords:
            fila_cabecera, columna_id = coords["ID_ANOMALIA"]

            # leemos y limpiamos filas vacías
            df_temp = pd.read_excel(ruta_xls, skiprows=fila_cabecera)
            df_temp = df_temp.dropna(how='all')

            # renombrado de columnas segun alias
            mapeo = {}
            for clave, (f, c) in coords.items():
                if c < len(df_temp.columns):
                    nombre_actual_columna = df_temp.columns[c]
                    mapeo[nombre_actual_columna] = clave

            df_temp = df_temp.rename(columns=mapeo)


            #si 'GRADO' no existe, lo buscamos manualmente
            if 'GRADO' not in df_temp.columns:
                col_grado_alternativa = next((c for c in df_temp.columns if 'grado' in str(c).lower()), None)
                if col_grado_alternativa:
                    df_temp = df_temp.rename(columns={col_grado_alternativa: 'GRADO'})


            # procesamos siempre
            if 'ID_ANOMALIA' in df_temp.columns:
                df_temp = df_temp.dropna(subset=['ID_ANOMALIA'])

                #volvemos añadir vano normal
                if 'VANO' in df_temp.columns:
                    df_temp['VANO_NORM'] = df_temp['VANO'].apply(normalizar_texto)

                if 'CIRCUITO' in df_temp.columns:
                    df_temp['CIRC_NORM'] = df_temp['CIRCUITO'].apply(normalizar_texto)

                nuevos_ids = []

                for _, fila in df_temp.iterrows():
                    id_original_excel = str(fila.iloc[columna_id]).split('.')[0]
                    id_nuevo = contador_global
                    nuevos_ids.append(id_nuevo)

                    builder_instance.product.image_map[id_nuevo] = {
                        "archivo_foto_original": id_original_excel,
                        "carpeta_origen": ruta_carpeta,
                        "hipotesis": nombre_hoja
                    }
                    contador_global += 1

                df_temp['Nº de anomalia'] = nuevos_ids
                df_temp['hipotesis_origen'] = nombre_hoja
                all_tables.append(df_temp)

    if all_tables:
        builder_instance.product.step1_raw_data = pd.concat(all_tables, ignore_index=True)
        builder_instance.product.data_dt = builder_instance.product.step1_raw_data

        print(f"Paso 1 terminado. {contador_global} anomalias procesadas.")


def select_worst_hypothesis_steep2(builder_instance, ruta_combinado):
    """
        Lee el Excel combinado y, para cada anomalía, busca qué hipótesis climática
        tiene el peor grado de gravedad y la distancia más corta.

        :param builder_instance: El objeto principal donde guardamos los datos del programa.
        :param ruta_combinado: La ruta al archivo Excel que tiene todas las hipótesis juntas.
        :return: No devuelve nada, guarda la tabla con las peores hipótesis dentro del 'builder_instance'.
    """
    try:
        info(f"Leyendo archivo de anomalías combinadas -> {ruta_combinado}")

        #leer el archivo combinado
        df_combinado = pd.read_excel(ruta_combinado)

        #identificar dinamicamente columnas de Grado (ignorando la generica)
        columnas_grado = [
            col for col in df_combinado.columns
            if 'grado' in str(col).lower() and str(col).lower().strip() != 'grado'
        ]

        #añadimos 'incumpli' a la lista para pillar "alincumplimiento"
        columnas_distancia = [
            col for col in df_combinado.columns
            if any(palabra in str(col).lower() for palabra in ['distancia', 'dist', 'incumpli'])
        ]

        col_long = next((c for c in df_combinado.columns if 'longitud' in str(c).lower()), None)
        filas_procesadas = []

        #iterar sobre todas las anomalías del excel combinado
        for index, fila in df_combinado.iterrows():
            max_grado = -1
            min_distancia = float('inf')
            peor_hipotesis = None


            for col_g in columnas_grado:
                #extraemos el nombre de la hipotesis
                hipotesis_sufijo = str(col_g).lower().replace('grado', '').strip()

                #buscamos la columna de distancia que termine EXACTAMENTE con esa hipótesis
                col_d_match = []
                for c in columnas_distancia:
                    #split() separa el texto por espacios y saltos de linea
                    palabras_columna = str(c).lower().split()
                    #comprobamos si la ultima palabra coincide exactamente
                    if palabras_columna and palabras_columna[-1] == hipotesis_sufijo:
                        col_d_match.append(c)

                if col_d_match:
                    col_d = col_d_match[0]  # Tomamos la coincidencia exacta

                    #extraer los valores y convertirlos a número
                    grado_actual = pd.to_numeric(fila[col_g], errors='coerce')
                    distancia_actual = pd.to_numeric(fila[col_d], errors='coerce')

                    #solo evaluamos si ambos datos son validos
                    if pd.notna(grado_actual) and pd.notna(distancia_actual):

                        #mayor grado, o a igualdad de grado, menor distancia.
                        if (grado_actual > max_grado) or (
                                grado_actual == max_grado and distancia_actual < min_distancia):
                            max_grado = grado_actual
                            min_distancia = distancia_actual

                            #guardamos el sufijo en mayusculas
                            peor_hipotesis = hipotesis_sufijo.upper()

            fila_dict = fila.to_dict()
            fila_dict['Hipotesis_Seleccionada'] = peor_hipotesis
            #solo guardamos los valores si encontramos alguna hipotesis valida
            fila_dict['Grado_Final'] = max_grado if max_grado != -1 else None
            fila_dict['Distancia_Final'] = min_distancia if min_distancia != float('inf') else None

            if col_long:
                fila_dict['Longitud_Vano_Final'] = pd.to_numeric(fila[col_long], errors='coerce')

            filas_procesadas.append(fila_dict)

        #convertir a dataframe y guardar en el objeto producto
        df_filtrado = pd.DataFrame(filas_procesadas)
        builder_instance.product.step2_filtered_data = df_filtrado

        info(f"Paso 2 completado. {len(df_filtrado)} anomalias filtradas por la peor hipotesis.")

    except Exception as e:
        error(f"Error critico en el Paso 2 procesando '{ruta_combinado}' -> {e}")
        raise

#cruzamos informacion de los combinados con los informes
def combine_information_steep3(builder_instance):
    """
    Cruza los datos de las peores hipótesis con la información cruda del principio
    para recuperar datos técnicos vitales como las coordenadas, el vano y el circuito.

    :param builder_instance: El objeto principal que ya contiene las tablas de los pasos 1 y 2.
    :return: No devuelve nada, guarda la tabla combinada y enriquecida en el 'builder_instance'.
    """
    info("Paso 3: Uniendo anomalias (Comparando...)")

    df_comb = builder_instance.product.step2_filtered_data
    df_inf = builder_instance.product.step1_raw_data

    # SOLUCIÓN CRÍTICA: Prevenir el error de fillna
    if 'GRADO' in df_inf.columns:
        df_inf['GRADO_NUM'] = pd.to_numeric(df_inf['GRADO'], errors='coerce').fillna(0).astype(int)
    else:
        df_inf['GRADO_NUM'] = 0

    col_vano_c = next((c for c in df_comb.columns if 'vano' in str(c).lower() and 'long' not in str(c).lower()), None)
    col_cir_c = next((c for c in df_comb.columns if 'circuito' in str(c).lower()), None)
    col_dist_inf = 'DISTANCIA' if 'DISTANCIA' in df_inf.columns else next(
        (c for c in df_inf.columns if 'dist' in str(c).lower()), None)

    if col_dist_inf:
        df_inf = df_inf.copy()
        df_inf['DIST_CALC'] = pd.to_numeric(df_inf[col_dist_inf].astype(str).str.replace(',', '.'), errors='coerce')

    filas_finales = []
    ids_asignados = set()
    col_desc_comb = next((c for c in df_comb.columns if 'descrip' in str(c).lower()), None)

    # Variables técnicas base
    cols_tecnicas = ['x', 'y', 'z', 'huso', 'REGLAMENTO', 'DIST_REGLA', 'APOYO_INI', 'APOYO_FIN', 'TENSION',
                     'LONGITUD_VANO', 'TIPO_ANO', 'CUMPLE', 'NOTAS', 'TEMP_MAX', 'FASE', 'DMR', 'VANO', 'CIRCUITO',
                     'DIST_APOYO']

    for idx, row_c in df_comb.iterrows():
        hipo_c = normalizar_texto(row_c.get('Hipotesis_Seleccionada'))
        vano_c = normalizar_texto(row_c.get(col_vano_c))
        dist_c_num = pd.to_numeric(str(row_c.get('Distancia_Final')).replace(',', '.'), errors='coerce')

        candidatos = df_inf[df_inf['hipotesis_origen'].apply(normalizar_texto) == hipo_c].copy()

        if not candidatos.empty and pd.notna(dist_c_num) and 'DIST_CALC' in candidatos.columns:
            candidatos['_diferencia'] = (candidatos['DIST_CALC'] - dist_c_num).abs()
            candidatos = candidatos.sort_values(by='_diferencia')

        mejor_match = None
        for _, cand in candidatos.iterrows():
            if cand['Nº de anomalia'] in ids_asignados: continue

            vano_i = cand.get('VANO_NORM', '')
            if vano_c and vano_i:
                coincide_vano = (vano_c == vano_i)
                if not coincide_vano and '-' in vano_c:
                    partes = vano_c.split('-')
                    if len(partes) == 2 and f"{partes[1]}-{partes[0]}" == vano_i: coincide_vano = True
                if not coincide_vano: continue

            cand_dist_num = cand.get('DIST_CALC')
            if pd.notna(dist_c_num) and pd.notna(cand_dist_num) and abs(dist_c_num - cand_dist_num) > 0.5: continue

            mejor_match = cand
            ids_asignados.add(cand['Nº de anomalia'])
            break

        fila_res = row_c.to_dict()

        if mejor_match is not None:
            fila_res['Nº de anomalia'] = mejor_match['Nº de anomalia']
            fila_res['ID_Original_Informe'] = mejor_match['ID_ANOMALIA']

            desc_final = fila_res.get(col_desc_comb, "") if col_desc_comb else ""
            if pd.isna(desc_final) or str(desc_final).strip() == "":
                col_desc_inf = next((c for c in mejor_match.keys() if
                                     any(p in str(c).lower() for p in ['descrip', 'observaci', 'texto', 'defecto'])),
                                    None)
                if col_desc_inf: desc_final = mejor_match[col_desc_inf]
            fila_res['Descripción'] = desc_final if pd.notna(desc_final) else ""

            # Rescatamos variables mapeadas por el diccionario
            for k in cols_tecnicas:
                val = mejor_match.get(k)
                if pd.notna(val) and str(val).strip() != "":
                    fila_res[k] = val


            # pillamos distancia apoyo
            if pd.isna(fila_res.get('DIST_APOYO')) or str(fila_res.get('DIST_APOYO')).strip() == "":
                for col_name in mejor_match.keys():
                    col_str = str(col_name).lower()
                    if 'distancia' in col_str and 'apoyo' in col_str:
                        val_apoyo = mejor_match[col_name]
                        if pd.notna(val_apoyo) and str(val_apoyo).strip() != "":
                            fila_res['DIST_APOYO'] = val_apoyo
                            break

            # pillamos Temperatura
            if pd.isna(fila_res.get('TEMP_MAX')) or str(fila_res.get('TEMP_MAX')).strip() == "":
                for col_name in mejor_match.keys():
                    col_str = str(col_name).lower()
                    if 'temp' in col_str:
                        val_temp = mejor_match[col_name]
                        if pd.notna(val_temp) and str(val_temp).strip() != "":
                            fila_res['TEMP_MAX'] = val_temp
                            break

        else:
            fila_res['Nº de anomalia'] = None
            fila_res['ID_Original_Informe'] = None
            desc_final = fila_res.get(col_desc_comb, "") if col_desc_comb else ""
            fila_res['Descripción'] = desc_final if pd.notna(desc_final) else ""

        # Vano y Circuito sin vacios
        if col_vano_c and (pd.isna(fila_res.get('VANO')) or str(fila_res.get('VANO')).strip() == ""):
            fila_res['VANO'] = row_c.get(col_vano_c)
        if col_cir_c and (pd.isna(fila_res.get('CIRCUITO')) or str(fila_res.get('CIRCUITO')).strip() == ""):
            fila_res['CIRCUITO'] = row_c.get(col_cir_c)

        filas_finales.append(fila_res)

    builder_instance.product.data_dt = pd.DataFrame(filas_finales)
    info(f"Paso 3 terminado. {len(ids_asignados)} anomalias vinculadas de forma exitosa")


def apply_porteman_steep4(builder_instance, ruta_porteman):
    """
        Compara la descripción de las anomalías con la anotacion porteman para
        asignarles automáticamente su código, la operación y si es Vegetación o Infraestructura.

        :param builder_instance: El objeto principal con los datos del paso 3.
        :param ruta_porteman: La ruta al archivo Excel que usamos como diccionario Porteman.
        :return: No devuelve nada, actualiza la tabla con los nuevos códigos y tipos.
        """
    try:
        info(f"Paso 4: Aplicando tabla Porteman desde -> {ruta_porteman}")

        if not ruta_porteman:
            error("No se ha proporcionado la ruta del fichero Porteman.")
            return

        df_porteman = pd.read_excel(ruta_porteman)
        df_porteman.columns = [str(c).strip() for c in df_porteman.columns]

        df_main = builder_instance.product.data_dt

        if df_main.empty:
            error("El DataFrame principal está vacío.")
            return

        col_desc_port = next((c for c in df_porteman.columns if 'cruzamiento' in str(c).lower()), None)
        col_cod_port = next((c for c in df_porteman.columns if 'cod' in str(c).lower()), None)

        # pillar columnas conjunto
        col_conjunto_port = next((c for c in df_porteman.columns if 'conjunto' in str(c).lower()), None)

        col_desc_main = 'Descripción'

        if not col_desc_port or not col_cod_port:
            error("No se encontraron las columnas necesarias en Porteman.")
            return

        df_porteman['_desc_norm'] = df_porteman[col_desc_port].apply(normalizar_texto)
        df_main['_desc_norm'] = df_main[col_desc_main].apply(normalizar_texto)

        cols_interes = [c for c in df_porteman.columns if c != '_desc_norm']
        porteman_dict = df_porteman.set_index('_desc_norm')[cols_interes].to_dict('index')

        filas_actualizadas = []
        palabras_vegetacion = ['veg', 'poda', 'tala', 'arbol', 'maleza', 'rama']

        for _, row in df_main.iterrows():
            row_dict = row.to_dict()
            desc_norm = row.get('_desc_norm', '')
            desc_original = str(row.get(col_desc_main, '')).lower()

            row_dict['Codigo'] = ""
            row_dict['Operacion'] = ""
            row_dict['Tipo_Anomalia'] = 'Infraestructura'
            row_dict['Conjunto'] = ""
            datos_encontrados = None

            if desc_norm in porteman_dict:
                datos_encontrados = porteman_dict[desc_norm]
            else:
                for texto_porteman, valores in porteman_dict.items():
                    if desc_norm and texto_porteman and (desc_norm in texto_porteman or texto_porteman in desc_norm):
                        datos_encontrados = valores
                        break

            if datos_encontrados:
                row_dict['Codigo'] = datos_encontrados.get(col_cod_port, "")

                # guardamos el value de conjunto
                if col_conjunto_port:
                    row_dict['Conjunto'] = datos_encontrados.get(col_conjunto_port, "")

                col_op = next((c for c in datos_encontrados.keys() if 'operaci' in str(c).lower()), None)
                if col_op:
                    row_dict['Operacion'] = datos_encontrados.get(col_op, "")

                valores_str = " ".join([str(v).lower() for v in datos_encontrados.values()])
                if any(p in desc_original for p in palabras_vegetacion) or any(
                        p in valores_str for p in palabras_vegetacion):
                    row_dict['Tipo_Anomalia'] = 'Vegetación'
            else:
                if any(p in desc_original for p in palabras_vegetacion):
                    row_dict['Tipo_Anomalia'] = 'Vegetación'

            filas_actualizadas.append(row_dict)

        builder_instance.product.data_dt = pd.DataFrame(filas_actualizadas)
        if '_desc_norm' in builder_instance.product.data_dt.columns:
            builder_instance.product.data_dt.drop(columns=['_desc_norm'], inplace=True)

        info(f"Paso 4 completado con éxito.")

    except Exception as e:
        error(f"Error en el Paso 4: {e}")
        raise


def link_vanos_info_steep5(builder_instance, ruta_vanos):
    """
        Busca el vano de cada anomalía en el archivo de vanos para
        averiguar su "Ubicación Técnica" (UT) exacta,
        separando si es un vano de infraestructura o de vegetación.

        :param builder_instance: El objeto principal con los datos del paso 4.
        :param ruta_vanos: La ruta al archivo excel que contiene la lista oficial de vanos y UTs.
        :return: No devuelve nada, añade la columna de ubicacion tecnica terminada a la tabla.
    """
    info("Paso 5: Vinculando información de vanos")
    try:
        df_p = builder_instance.product.data_dt
        df_vanos = pd.read_excel(ruta_vanos)

        #buscamos las columnas de forma dinamica
        col_vano_anomalias = next(
            (c for c in df_p.columns if 'vano' in str(c).lower() and 'long' not in str(c).lower()), None)

        if not col_vano_anomalias:
            df_p['Ubicación Técnica'] = ""
            builder_instance.product.data_dt = df_p
            return

        df_vanos.columns = [str(c).lower().strip() for c in df_vanos.columns]
        col_denominacion = next((col for col in df_vanos.columns if 'denominaci' in col), None)
        col_ubicacion = next((col for col in df_vanos.columns if 'ubicaci' in col), None)

        if not col_denominacion or not col_ubicacion:
            #a nivel de columna entera si no hay maestro
            df_p['Ubicación Técnica'] = df_p[col_vano_anomalias]
            builder_instance.product.data_dt = df_p
            return

        def obtener_ut(fila):
            vano_anomalia = str(fila.get(col_vano_anomalias, '')).replace(" ", "").strip()
            # aseguramos que pillamos bien el tipo de anomalía
            tipo_anomalia = str(fila.get('Tipo_Anomalia', '')).lower().strip()
            es_vegetacion = 'vegetaci' in tipo_anomalia

            if not vano_anomalia or vano_anomalia == 'nan':
                return ""

            # i  nvertimos el vano (de 571-572 a 572-571) por si en el maestro esta al reves
            partes = vano_anomalia.split('-')
            if len(partes) == 2:
                vano_inverso = f"{partes[1]}-{partes[0]}"
            else:
                vano_inverso = vano_anomalia

            serie_denominacion = df_vanos[col_denominacion].astype(str)
            serie_ubicacion = df_vanos[col_ubicacion].astype(str)

            # buscamos coincidencias con guiones para ser exactos
            vano_busqueda_1 = f"-{vano_anomalia}-"
            vano_busqueda_2 = f"-{vano_inverso}-"

            mask = serie_denominacion.str.contains(vano_busqueda_1, regex=False, na=False) | \
                   serie_denominacion.str.contains(vano_busqueda_2, regex=False, na=False)

            coincidencias = df_vanos[mask]

            # si no hay con guiones, probamos el vano en bruto
            if coincidencias.empty:
                mask_fallback = serie_denominacion.str.contains(vano_anomalia, regex=False, na=False) | \
                                serie_denominacion.str.contains(vano_inverso, regex=False, na=False)
                coincidencias = df_vanos[mask_fallback]

            # si no se encuentra en el maestro, devolvemos el vano original
            if coincidencias.empty:
                return vano_anomalia

            # reglas para v y vf
            valores_ut = coincidencias[col_ubicacion].astype(str).str.strip()

            #si es de vegetacion empezara por VF, sino empezara por v
            if es_vegetacion:

                match = valores_ut[valores_ut.str.startswith('VF', na=False)]
                if not match.empty:
                    return match.iloc[0]
                else:
                    # encontrado pero sin VF, devolvemos vano
                    return vano_anomalia
            else:

                match = valores_ut[
                    valores_ut.str.startswith('V', na=False) & ~valores_ut.str.startswith('VF', na=False)]
                if not match.empty:
                    return match.iloc[0]
                else:
                    # encontrado pero sin V, devolvemos vano
                    return vano_anomalia

        #guardamos
        df_p['Ubicación Técnica'] = df_p.apply(obtener_ut, axis=1)
        builder_instance.product.data_dt = df_p
        info("Paso 5 completado con éxito.")

    except Exception as e:
        error(f"Error en el Paso 5: {e}")
        # En caso de error fatal, rellenar con el vano
        col_vano = next((c for c in builder_instance.product.data_dt.columns if
                         'vano' in str(c).lower() and 'long' not in str(c).lower()), None)
        if col_vano:
            builder_instance.product.data_dt['Ubicación Técnica'] = builder_instance.product.data_dt[col_vano]


def export_final_data_steep6(builder_instance, carpeta_salida, orden_trabajo):
    """
        Coge toda la información procesada y crea los excels finales,
        aplicando los estilos y copiando las fotos a sus carpetas.

        :param builder_instance: El objeto principal con la tabla final, lista para imprimir.
        :param carpeta_salida: La ruta donde se van a guardar los excels y las carpetas de fotos.
        :param orden_trabajo: El número de orden de trabajo para ponerlo en el nombre de los archivos.
        :return: No devuelve nada, genera los archivos y carpetas.
    """
    info("Paso 6: Exportando archivos xls (Factory Builder)")
    try:
        df = builder_instance.product.data_dt.copy()
        image_map = builder_instance.product.image_map

        if df.empty or not carpeta_salida:
            error("Faltan datos o carpeta de salida.")
            return

        fecha_actual = datetime.now().strftime("%d/%m/%Y")
        prefijo_circuito = next(
            (os.path.basename(ruta).split('_')[0] for ruta in builder_instance.product.source_paths.values() if
             os.path.basename(ruta).split('_')[0].isdigit()), "")
        col_circ = next((c for c in df.columns if 'circuito' in str(c).lower()), None)

        if not prefijo_circuito:
            prefijo_circuito = ''.join(filter(str.isdigit, str(df[col_circ].iloc[0])))[:4] if col_circ else "0000"

        df['ID'] = df.apply(lambda fila: (
                f"1{prefijo_circuito}" + str(fila.get('ID_Original_Informe', '')).split('.')[0].zfill(
            10 - len(f"1{prefijo_circuito}"))) if str(fila.get('ID_Original_Informe', '')).split('.')[
            0].isdigit() else str(fila.get('ID_Original_Informe', '')).split('.')[0], axis=1)

        col_ut = 'Ubicación Técnica'
        col_vano = next((c for c in df.columns if 'vano' in str(c).lower() and 'long' not in str(c).lower()), None)
        df[col_ut] = df.apply(
            lambda fila: str(fila.get(col_vano, '')).strip() if not str(fila.get(col_ut, '')).strip() or str(
                fila.get(col_ut, '')).strip().lower() in ['nan', 'none'] else str(fila.get(col_ut, '')).strip(), axis=1)

        #redondeamos a 2 decimales
        cols_a_redondear = ['x', 'y', 'z', 'LONGITUD_VANO', 'DIST_REGLA', 'Distancia_Final', 'CUMPLE']
        for col_red in cols_a_redondear:
            if col_red in df.columns:
                df[col_red] = pd.to_numeric(df[col_red], errors='coerce').round(2)

        # ordenamos de menor a mayor las anomalias
        df = df.sort_values(by='ID', ascending=True)

        nombre_circuito = str(df[col_circ].iloc[0]).strip() if col_circ else "CIRCUITO"

        tipos_salida = {
            'Infraestructura': [f"{nombre_circuito}_AnomaliasInfraestructuras",
                                f"{nombre_circuito}_AnomaliasInfraestructuras_ANEXOX_LITE",
                                f"{nombre_circuito}_AnomaliasInfraestructuras2026"],
            'Vegetación': [f"{nombre_circuito}_AnomaliasVegetacion",
                           f"{nombre_circuito}_AnomaliasVegetacion_ANEXOX_LITE"]
        }

        for tipo, nombres_archivos in tipos_salida.items():
            df_tipo = df[df['Tipo_Anomalia'] == tipo].copy()
            if df_tipo.empty: continue

            ruta_fotos_tipo = os.path.join(carpeta_salida, tipo)
            os.makedirs(ruta_fotos_tipo, exist_ok=True)
            df_tipo['Evidencia_Grafica_Ruta'] = ""

            #proceso para asociar ids a imagenes
            for index, fila in df_tipo.iterrows():
                id_interno = fila.get('Nº de anomalia')
                if pd.notna(id_interno) and id_interno in image_map:
                    datos_foto = image_map[id_interno]
                    ruta_origen_fotos = os.path.join(datos_foto["carpeta_origen"], "pictures")
                    id_foto_original = str(datos_foto["archivo_foto_original"]).strip()
                    if os.path.exists(ruta_origen_fotos):
                        archivo_real = next((f for f in os.listdir(ruta_origen_fotos) if
                                             f.startswith(id_foto_original) and f.lower().endswith(
                                                 ('.jpg', '.jpeg', '.png'))), None)
                        if archivo_real:
                            nuevo_nombre = f"{str(fila.get(col_ut, 'UT')).replace('/', '-')}_{str(fila.get('Codigo', 'NA')).replace('/', '-')}_{str(fila.get('ID', id_interno))}{os.path.splitext(archivo_real)[1]}"
                            shutil.copy2(os.path.join(ruta_origen_fotos, archivo_real),
                                         os.path.join(ruta_fotos_tipo, nuevo_nombre))
                            df_tipo.at[index, 'Evidencia_Grafica_Ruta'] = f"{tipo}/{nuevo_nombre}"

            for nombre_base in nombres_archivos:
                ruta_temp_xlsx = os.path.join(carpeta_salida,
                                              f"{orden_trabajo}_{nombre_base}.xlsx" if orden_trabajo else f"{nombre_base}.xlsx")
                ruta_final_xls = os.path.join(carpeta_salida,
                                              f"{orden_trabajo}_{nombre_base}.xls" if orden_trabajo else f"{nombre_base}.xls")

                if "LITE" in nombre_base:
                    cols_lite = ['DMR', 'huso', 'x', 'y', 'z', 'ID', 'Grado_Final', 'CIRCUITO', 'VANO', 'APOYO_INI',
                                 'APOYO_FIN', 'TENSION', 'LONGITUD_VANO', 'TIPO_ANO', 'REGLAMENTO', 'DIST_REGLA',
                                 'Distancia_Final', 'CUMPLE', 'NOTAS']
                    df_lite = df_tipo.reindex(columns=cols_lite)
                    df_lite.to_excel(ruta_temp_xlsx, index=False, startrow=3, header=False, engine='openpyxl')

                    wb = load_workbook(ruta_temp_xlsx)
                    ws = wb.active

                    ama = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                        bottom=Side(style='thin'))

                    alin_vertical = Alignment(horizontal='center', vertical='bottom', textRotation=90, wrapText=True)
                    alin_horizontal = Alignment(horizontal='center', vertical='center', wrapText=True)
                    alin_datos = Alignment(horizontal='center', vertical='center')


                    alin_coordenadas = Alignment(horizontal='center', vertical='center', textRotation=90, wrapText=True)

                    ws.merge_cells('C1:E1');
                    ws['C1'] = "UTM de las Anomalías";
                    ws['C1'].alignment = alin_horizontal
                    ws.merge_cells('C2:E2');
                    ws['C2'] = "Coordenadas";
                    ws['C2'].alignment = alin_coordenadas
                    ws['C3'] = "UTM (X)";
                    ws['C3'].alignment = alin_horizontal
                    ws['D3'] = "UTM (Y)";
                    ws['D3'].alignment = alin_horizontal
                    ws['E3'] = "UTM (Z)";
                    ws['E3'].alignment = alin_horizontal

                    titulos_verticales = {
                        1: "DMR", 2: "Huso", 6: "Nº de anomalía", 7: "Grado de anomalía",
                        8: "Nemónico circuito", 9: "Vano", 10: "Nº apoyo (a) inicial",
                        11: "Nº apoyo (a) final", 12: "Tensión (kV)", 13: "Longitud vano",
                        14: "Tipo de anomalía", 15: "Se aplica el reglamento", 16: "Distancia de reglamento",
                        17: "Mínima distancia absoluta de (h) a la trayectoria del cable",
                        18: "Cumple/incumple por +/- (m)", 19: "Notas"
                    }

                    for col_idx, texto in titulos_verticales.items():
                        letra = get_column_letter(col_idx)
                        ws.merge_cells(f'{letra}1:{letra}3')
                        ws[f'{letra}1'] = texto
                        ws[f'{letra}1'].alignment = alin_vertical

                    # Pintar cabecera
                    for fila in range(1, 4):
                        for col in range(1, 20):
                            celda = ws.cell(row=fila, column=col)
                            celda.fill = ama
                            celda.border = borde_fino

                    # Centrar los datos y poner cuadrícula a las filas de abajo
                    for r_idx in range(4, ws.max_row + 1):
                        for c_idx in range(1, 20):
                            celda = ws.cell(row=r_idx, column=c_idx)
                            celda.border = borde_fino
                            celda.alignment = alin_datos

                    ws.row_dimensions[1].height = 25
                    ws.row_dimensions[2].height = 140
                    ws.row_dimensions[3].height = 20

                    for i in range(1, 20):
                        if i in [3, 4, 5]:
                            ws.column_dimensions[get_column_letter(i)].width = 14
                        else:
                            ws.column_dimensions[get_column_letter(i)].width = 7

                    wb.save(ruta_temp_xlsx)
                    wb.close()


                elif "2026" in nombre_base:

                    df_ext = pd.DataFrame()

                    df_ext['Nº Anomalia SAP'] = df_tipo['ID']
                    df_ext['Ubicación técnica'] = df_tipo['Ubicación Técnica']
                    df_ext['Conjunto'] = df_tipo.get('Conjunto', "")
                    df_ext['Texto Ampliado'] = ""
                    df_ext['Repercusión'] = df_tipo.get('Grado_Final', 0)
                    df_ext['Reglamento de Aplicación'] = df_tipo.get('REGLAMENTO', "")
                    df_ext['Temperatura max de diseño de la linea'] = df_tipo.get('TEMP_MAX', "")
                    codigos = df_tipo.get('Codigo', "").fillna("")
                    descripciones = df_tipo.get('Descripción', "").fillna("")
                    df_ext['Código caracteristico - Descripción'] = codigos.astype(str) + " - " + descripciones.astype(str)
                    df_ext['Distancia al primer apoyo del vano'] = df_tipo.get('DIST_APOYO', "")
                    df_ext['Fase/CT en la que se encuentra el incumplimiento'] = 3
                    df_ext['X'] = df_tipo.get('x', "")
                    df_ext['Y'] = df_tipo.get('y', "")
                    df_ext['HUSO'] = df_tipo.get('huso', "")
                    df_ext['Hipótesis climática de incumplimiento'] = df_tipo['Hipotesis_Seleccionada']
                    df_ext['Distancia reglamentaria considerada'] = df_tipo.get('DIST_REGLA', "")
                    df_ext['Distancia en el punto de incumplimiento'] = df_tipo['Distancia_Final']
                    df_ext.to_excel(ruta_temp_xlsx, index=False, engine='openpyxl')

                    wb = load_workbook(ruta_temp_xlsx);
                    ws = wb.active
                    ws.cell(row=1, column=17, value="Evidencia Gráfica")
                    link_font = Font(color="0000FF", underline="single")
                    for r_idx, evi in enumerate(df_tipo['Evidencia_Grafica_Ruta'], 2):

                        if evi:
                            c = ws.cell(row=r_idx, column=17, value="Imagen")
                            c.hyperlink = evi;
                            c.font = link_font

                    for i in range(1, 18):
                        ws.column_dimensions[get_column_letter(i)].width = 22

                    wb.save(ruta_temp_xlsx);
                    wb.close()

                else:

                    df_n = pd.DataFrame()
                    df_n['ID'] = df_tipo['ID']
                    df_n['UT'] = df_tipo['Ubicación Técnica']
                    df_n['DESCRIPCIÓN'] = ""
                    df_n['TEXTO AMPLIADO'] = ""
                    df_n['REPERCUSIÓN'] = df_tipo.get('Grado_Final', 0)
                    df_n['FECHA INICIO AVISO'] = fecha_actual
                    df_n['ORDEN ORIGEN'] = str(orden_trabajo)
                    df_n['FECHA INICIO AVERÍA'] = fecha_actual
                    df_n['OPERACIÓN'] = df_tipo.get('Operacion', "")

                    df_n.to_excel(ruta_temp_xlsx, index=False, engine='openpyxl')
                    wb = load_workbook(ruta_temp_xlsx);
                    ws = wb.active
                    ws.cell(row=1, column=10, value="EVIDENCIA GRÁFICA")
                    link_font = Font(color="0000FF", underline="single")
                    for r_idx, evi in enumerate(df_tipo['Evidencia_Grafica_Ruta'], 2):
                        if evi:
                            c = ws.cell(row=r_idx, column=10, value="Imagen")
                            c.hyperlink = evi;
                            c.font = link_font
                    for i in range(1, 11): ws.column_dimensions[get_column_letter(i)].width = 22
                    wb.save(ruta_temp_xlsx);
                    wb.close()

                if os.path.exists(ruta_final_xls): os.remove(ruta_final_xls)
                os.rename(ruta_temp_xlsx, ruta_final_xls)
                info(f"Creado archivo: {nombre_base}")

    except Exception as e:
        error(f"Fallo al exportar: {e}")
        raise